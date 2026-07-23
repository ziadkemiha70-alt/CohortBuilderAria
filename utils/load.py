# -*- coding: utf-8 -*-
"""Lecture robuste des fichiers CSV/Excel/ZIP et streaming formulaire."""

import contextlib
import csv
import io
import re
import zipfile
from typing import Any, Iterable, List, Optional, Tuple

import pandas as pd
import streamlit as st

from utils.clean import nonempty_mask, normalize_pt_key

# Force l'invalidation des caches Streamlit lorsque la logique de lecture CSV change.
# Important après la correction des guillemets doubles dans les exports ARIA.
CSV_PARSER_CACHE_VERSION = "2026_06_11_fix_csv_quotes_v2"


def _detect_text_encoding_from_bytes(sample: bytes) -> str:
    """Détecte un encodage sans charger tout le fichier."""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            sample.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin1"


def _guess_csv_dialect(sample_text: str) -> csv.Dialect:
    """Détecte le séparateur CSV avec une gestion prudente des guillemets.

    Important : ``csv.Sniffer`` peut parfois proposer ``doublequote=False`` sur
    les exports ARIA contenant des intitulés comme ``si ""oui"": localisation,
    lieu``. Dans ce cas, le lecteur découpe à tort la virgule située dans le
    libellé de colonne, ajoute une colonne fantôme dans l'en-tête, puis décale
    toutes les colonnes suivantes. C'est typiquement ce qui faisait lire
    ``[SEC Préparation NC SEIN]`` à la place de
    ``Reprise d'une activité sexuelle``.

    On utilise donc Sniffer uniquement pour le séparateur, puis on force un
    dialecte standard avec guillemets doubles correctement interprétés.
    """
    delimiters = [",", ";", "\t", "|"]
    try:
        sniffed = csv.Sniffer().sniff(sample_text, delimiters="".join(delimiters))
        delimiter = getattr(sniffed, "delimiter", ",") or ","
    except Exception:
        first_line = sample_text.splitlines()[0] if sample_text.splitlines() else ""
        counts = {d: first_line.count(d) for d in delimiters}
        delimiter = max(counts, key=counts.get) if counts else ","

    class RobustDialect(csv.excel):
        pass

    RobustDialect.delimiter = delimiter
    RobustDialect.quotechar = '"'
    RobustDialect.doublequote = True
    RobustDialect.escapechar = None
    RobustDialect.quoting = csv.QUOTE_MINIMAL
    RobustDialect.skipinitialspace = False
    RobustDialect.lineterminator = "\n"
    return RobustDialect


@contextlib.contextmanager
def _open_csv_text_stream(uploaded_file):
    """Ouvre un CSV uploadé sans passer par pandas ni dupliquer le fichier entier.

    Pour éviter l'erreur `Error tokenizing data: out of memory`, le formulaire
    est lu avec le module standard `csv`, qui itère ligne par ligne et respecte
    les guillemets/multilignes sans construire de gros tableaux intermédiaires.
    """
    name = uploaded_file.name.lower()
    if name.endswith(".zip"):
        uploaded_file.seek(0)
        zf = zipfile.ZipFile(uploaded_file)
        csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
        if not csv_names:
            zf.close()
            raise ValueError("Le ZIP ne contient aucun CSV.")
        bio = zf.open(csv_names[0], "r")
        sample = bio.read(65536)
        enc = _detect_text_encoding_from_bytes(sample)
        bio.close()
        bio = zf.open(csv_names[0], "r")
        text = io.TextIOWrapper(bio, encoding=enc, newline="")
        try:
            yield text
        finally:
            try:
                text.detach()
            except Exception:
                pass
            bio.close()
            zf.close()
    else:
        uploaded_file.seek(0)
        sample = uploaded_file.read(65536)
        enc = _detect_text_encoding_from_bytes(sample)
        uploaded_file.seek(0)
        text = io.TextIOWrapper(uploaded_file, encoding=enc, newline="")
        try:
            yield text
        finally:
            # Important : ne pas fermer l'objet Streamlit, on le détache juste.
            try:
                text.detach()
            except Exception:
                pass
            try:
                uploaded_file.seek(0)
            except Exception:
                pass


def _csv_header_stdlib(uploaded_file) -> List[str]:
    with _open_csv_text_stream(uploaded_file) as text:
        sample = text.read(65536)
        dialect = _guess_csv_dialect(sample)
        text.seek(0)
        reader = csv.reader(text, dialect)
        try:
            header = next(reader)
        except StopIteration:
            return []
        return [str(h).strip() for h in header]


def _iter_csv_chunks_from_upload(
    uploaded_file, usecols: Optional[List[str]] = None, chunksize: int = 1000
):
    """Itère sur un CSV uploadé en petits morceaux sans pandas.read_csv.

    Cette version évite complètement le parseur C de pandas, responsable de
    l'erreur `Error tokenizing data C error out of memory` sur les très grands
    formulaires ARIA. Elle garde toutes les fonctionnalités : sélection de
    colonnes, calcul des stats, création du formulaire long et export final.
    """
    name = uploaded_file.name.lower()
    if name.endswith((".xlsx", ".xls")):
        df = read_uploaded_table(uploaded_file, usecols=usecols)
        yield protect_binary_checkbox_codes(df, columns=usecols)
        return

    with _open_csv_text_stream(uploaded_file) as text:
        sample = text.read(65536)
        dialect = _guess_csv_dialect(sample)
        text.seek(0)
        reader = csv.reader(text, dialect)
        try:
            header = [str(h).strip() for h in next(reader)]
        except StopIteration:
            return

        if usecols is None:
            selected_cols = header
        else:
            selected_cols = [c for c in usecols if c in header]
        selected_idx = [header.index(c) for c in selected_cols]

        rows = []
        for row in reader:
            # On ne saute aucune ligne. Les lignes plus courtes sont complétées.
            # Les champs surnuméraires n'ont pas de nom de colonne, donc ils ne
            # peuvent pas être sélectionnés par `usecols`.
            if len(row) < len(header):
                row = row + [""] * (len(header) - len(row))
            rows.append([row[i] if i < len(row) else "" for i in selected_idx])
            if len(rows) >= chunksize:
                chunk = pd.DataFrame(rows, columns=selected_cols, dtype=object)
                yield protect_binary_checkbox_codes(chunk, columns=selected_cols)
                rows = []
        if rows:
            chunk = pd.DataFrame(rows, columns=selected_cols, dtype=object)
            yield protect_binary_checkbox_codes(chunk, columns=selected_cols)


def _excel_zero_format_width(number_format: Any) -> Optional[int]:
    """Retourne la largeur d'un format Excel du type 00000.

    Certains formulaires contiennent des codes binaires de cases à cocher
    (ex. 00100). Si Excel stocke la cellule comme nombre mais lui applique un
    format 00000, pandas lit 100. Cette fonction permet de reconstruire la
    valeur affichée par Excel au moment de la lecture.
    """
    fmt = str(number_format or "").split(";")[0].strip()
    if re.fullmatch(r"0{2,}", fmt):
        return len(fmt)
    return None


def _format_excel_cell_preserving_ids(cell: Any) -> Any:
    value = cell.value
    if value is None:
        return ""

    width = _excel_zero_format_width(getattr(cell, "number_format", ""))
    if width and isinstance(value, (int, float)):
        try:
            fval = float(value)
            if fval.is_integer():
                return str(int(fval)).zfill(width)
        except Exception:
            pass

    return value


def _normalize_binary_checkbox_code_value(value: Any, width: int) -> Any:
    """Protège les codes binaires de formulaires contre les conversions numériques.

    Exemples : 100 -> 00100 si une autre valeur de la colonne indique une
    largeur de 5 ; 1 -> 00001. Les valeurs non binaires sont laissées inchangées.
    """
    if pd.isna(value):
        return value
    raw = str(value).strip()
    if raw.lower() in {"", "nan", "<na>", "none", "na"}:
        return value
    if re.fullmatch(r"[01]+(?:\.0+)?", raw):
        raw = re.sub(r"\.0+$", "", raw)
        return raw.zfill(width) if len(raw) < width else raw
    return value


def protect_binary_checkbox_codes(
    df: pd.DataFrame, columns: Optional[List[str]] = None
) -> pd.DataFrame:
    """Préserve les colonnes de formulaire codées comme identifiants binaires.

    Les toxicités/choix multiples peuvent être encodés sous la forme 00100,
    01001, etc. Ces valeurs ne doivent pas être traitées comme des nombres,
    sinon les zéros à gauche disparaissent. La détection reste prudente : une
    colonne est corrigée uniquement si ses valeurs non vides sont composées de
    0/1 et qu'au moins une valeur indique un code à plusieurs positions.
    """
    if df is None or df.empty:
        return df

    out = df.copy()
    target_cols = [c for c in (columns or list(out.columns)) if c in out.columns]
    for col in target_cols:
        non_empty = out[col][nonempty_mask(out[col])]
        if non_empty.empty:
            continue

        as_text = (
            non_empty.astype(str).str.strip().str.replace(r"\.0+$", "", regex=True)
        )
        binary_mask = as_text.str.fullmatch(r"[01]+", na=False)
        if not bool(binary_mask.all()):
            continue

        max_width = int(as_text.str.len().max()) if len(as_text) else 0
        has_multibit_code = max_width >= 2
        has_visible_leading_zero = bool(as_text.str.match(r"^0[01]+$", na=False).any())
        if not (has_multibit_code or has_visible_leading_zero):
            continue

        # On utilise la plus grande largeur observée. Si Excel a déjà supprimé
        # tous les zéros et qu'aucune cellule ne porte plus la largeur originale,
        # cette largeur ne peut pas être devinée de façon fiable.
        width = max(max_width, 2)
        out[col] = out[col].map(
            lambda v, w=width: _normalize_binary_checkbox_code_value(v, w)
        )
    return out


def _read_excel_preserve_ids(
    uploaded_file, usecols: Optional[List[str]] = None, nrows: Optional[int] = None
) -> pd.DataFrame:
    """Lit un Excel en préservant les cellules textuelles et les formats 00000."""
    raw = uploaded_file.getvalue()
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows()
        try:
            header_cells = next(rows_iter)
        except StopIteration:
            return pd.DataFrame()

        header = [
            str(c.value).strip() if c.value is not None else "" for c in header_cells
        ]
        if usecols is None:
            selected_cols = header
        else:
            selected_cols = [c for c in usecols if c in header]
        selected_idx = [header.index(c) for c in selected_cols]

        rows = []
        for i, row in enumerate(rows_iter):
            if nrows is not None and i >= nrows:
                break
            values = []
            for idx in selected_idx:
                cell = row[idx] if idx < len(row) else None
                values.append(
                    _format_excel_cell_preserving_ids(cell) if cell is not None else ""
                )
            rows.append(values)
        wb.close()
        return pd.DataFrame(rows, columns=selected_cols, dtype=object)
    except Exception:
        return pd.read_excel(
            io.BytesIO(raw),
            dtype=object,
            keep_default_na=False,
            na_filter=False,
            usecols=usecols,
            nrows=nrows,
        )


@st.cache_data(show_spinner=False)
def read_uploaded_columns(
    uploaded_file, cache_version: str = CSV_PARSER_CACHE_VERSION
) -> List[str]:
    """Lit uniquement l'en-tête du fichier, sans charger les lignes."""
    name = uploaded_file.name.lower()
    if name.endswith((".xlsx", ".xls")):
        raw = uploaded_file.getvalue()
        return pd.read_excel(io.BytesIO(raw), nrows=0).columns.tolist()
    return _csv_header_stdlib(uploaded_file)


def read_csv_selected_stdlib(
    uploaded_file,
    usecols: Optional[List[str]] = None,
    nrows: Optional[int] = None,
    chunksize: int = 5000,
) -> pd.DataFrame:
    """Construit un DataFrame à partir du CSV standard, uniquement pour les colonnes utiles."""
    chunks = []
    remaining = nrows
    for chunk in _iter_csv_chunks_from_upload(
        uploaded_file, usecols=usecols, chunksize=chunksize
    ):
        if remaining is not None:
            if remaining <= 0:
                break
            if len(chunk) > remaining:
                chunk = chunk.iloc[:remaining].copy()
            remaining -= len(chunk)
        chunks.append(chunk)
        if remaining is not None and remaining <= 0:
            break
    if not chunks:
        cols = usecols or read_uploaded_columns(uploaded_file)
        return pd.DataFrame(columns=cols)
    return pd.concat(chunks, ignore_index=True, sort=False)


def _try_read_csv(uploaded_file, **kwargs) -> pd.DataFrame:
    """Compatibilité : remplace pandas.read_csv par le lecteur standard mémoire-sûr.

    Les kwargs pandas non pertinents sont ignorés volontairement.
    """
    usecols = kwargs.get("usecols")
    nrows = kwargs.get("nrows")
    return read_csv_selected_stdlib(uploaded_file, usecols=usecols, nrows=nrows)


@st.cache_data(show_spinner=False)
def read_uploaded_table(
    uploaded_file,
    usecols: Optional[List[str]] = None,
    nrows: Optional[int] = None,
    cache_version: str = CSV_PARSER_CACHE_VERSION,
) -> pd.DataFrame:
    """Lecture robuste avec possibilité de ne charger qu'un sous-ensemble."""
    name = uploaded_file.name.lower()
    if name.endswith((".xlsx", ".xls")):
        df = _read_excel_preserve_ids(uploaded_file, usecols=usecols, nrows=nrows)
        return protect_binary_checkbox_codes(df, columns=usecols)
    df = read_csv_selected_stdlib(uploaded_file, usecols=usecols, nrows=nrows)
    return protect_binary_checkbox_codes(df, columns=usecols)


@st.cache_data(show_spinner=False)
def load_form_meta_only(
    uploaded_file, cache_version: str = CSV_PARSER_CACHE_VERSION
) -> Tuple[pd.DataFrame, List[str]]:
    """Charge seulement pt_id/date_event et renvoie aussi toutes les colonnes.

    Aucune information n'est perdue : les autres colonnes sont relues à la
    demande en streaming lors du calcul des stats et de l'export.
    """
    cols = read_uploaded_columns(uploaded_file)
    needed = [c for c in ["pt_id", "date_event"] if c in cols]
    missing = [c for c in ["pt_id", "date_event"] if c not in cols]
    if missing:
        raise ValueError(f"Colonnes manquantes dans le formulaire : {missing}")
    meta = read_uploaded_table(uploaded_file, usecols=needed)
    return meta, cols


@st.cache_data(show_spinner=False)
def compute_form_column_stats_streaming(
    uploaded_file,
    candidate_cols: List[str],
    cohort_keys: Iterable[Any],
    chunksize: int = 2500,
) -> pd.DataFrame:
    """Calcule Valeurs cohorte / Patients cohorte sans charger tout le formulaire."""
    cohort_keys = {str(k) for k in cohort_keys if pd.notna(k)}
    value_counts = {c: 0 for c in candidate_cols}
    patient_sets = {c: set() for c in candidate_cols}
    if not candidate_cols or not cohort_keys:
        return pd.DataFrame(
            {
                "Colonne formulaire": candidate_cols,
                "Valeurs cohorte": 0,
                "Patients cohorte": 0,
            }
        )

    usecols = ["pt_id"] + candidate_cols
    for chunk in _iter_csv_chunks_from_upload(
        uploaded_file, usecols=usecols, chunksize=chunksize
    ):
        if "pt_id" not in chunk.columns:
            continue
        chunk = chunk.copy()
        chunk["_pt_join_key"] = normalize_pt_key(chunk["pt_id"])
        chunk = chunk[chunk["_pt_join_key"].astype(str).isin(cohort_keys)]
        if chunk.empty:
            continue
        for c in candidate_cols:
            if c not in chunk.columns:
                continue
            m = nonempty_mask(chunk[c])
            n = int(m.sum())
            if n:
                value_counts[c] += n
                patient_sets[c].update(
                    chunk.loc[m, "_pt_join_key"].dropna().astype(str).unique().tolist()
                )
    return pd.DataFrame(
        {
            "Colonne formulaire": candidate_cols,
            "Valeurs cohorte": [value_counts[c] for c in candidate_cols],
            "Patients cohorte": [len(patient_sets[c]) for c in candidate_cols],
        }
    )


@st.cache_data(show_spinner=False)
def prepare_forms_long_streaming(
    uploaded_file,
    patient_base: pd.DataFrame,
    selected_cols: List[str],
    deduplicate: bool,
    chunksize: int = 2500,
    cache_version: str = CSV_PARSER_CACHE_VERSION,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Construit le formulaire long en streaming sur les colonnes sélectionnées."""
    if not selected_cols:
        return pd.DataFrame(), pd.DataFrame()
    cohort_keys = {str(k) for k in patient_base["_pt_join_key"].dropna().astype(str)}
    usecols = ["pt_id", "date_event"] + selected_cols
    frames: List[pd.DataFrame] = []

    for chunk in _iter_csv_chunks_from_upload(
        uploaded_file, usecols=usecols, chunksize=chunksize
    ):
        if "pt_id" not in chunk.columns or "date_event" not in chunk.columns:
            continue
        chunk = chunk.copy()
        chunk["_pt_join_key"] = normalize_pt_key(chunk["pt_id"])
        chunk = chunk[chunk["_pt_join_key"].astype(str).isin(cohort_keys)]
        if chunk.empty:
            continue
        value_vars = [c for c in selected_cols if c in chunk.columns]
        if not value_vars:
            continue
        chunk = protect_binary_checkbox_codes(chunk, columns=value_vars)
        long = chunk.melt(
            id_vars=["pt_id", "_pt_join_key", "date_event"],
            value_vars=value_vars,
            var_name="item",
            value_name="Donnee",
        )
        long = long[nonempty_mask(long["Donnee"])].copy()
        if not long.empty:
            frames.append(long)

    if not frames:
        return (
            pd.DataFrame(
                columns=[
                    "pt_id",
                    "_pt_join_key",
                    "date_event",
                    "item",
                    "Donnee",
                    "DateHeure",
                    "startD",
                    "endD",
                    "delai_startD_jours",
                    "delai_endD_jours",
                ]
            ),
            pd.DataFrame(),
        )

    long = pd.concat(frames, ignore_index=True, sort=False)
    long["DateHeure"] = pd.to_datetime(long["date_event"], errors="coerce")
    long = long.merge(
        patient_base[["_pt_join_key", "startD", "endD"]], on="_pt_join_key", how="left"
    )
    long["delai_startD_jours"] = (long["DateHeure"] - long["startD"]).dt.days
    long["delai_endD_jours"] = (long["DateHeure"] - long["endD"]).dt.days
    dup_cols = ["_pt_join_key", "item", "DateHeure", "Donnee"]
    duplicates = long[long.duplicated(dup_cols, keep=False)].copy()
    if deduplicate:
        long = long.drop_duplicates(dup_cols, keep="first").copy()
    return (
        long.sort_values(["_pt_join_key", "item", "DateHeure"], kind="stable"),
        duplicates,
    )


def load_ethos_table(uploaded_file) -> pd.DataFrame:
    """Charge le fichier ETHOS optionnel.

    Il est lu avec la même logique robuste que traitement_patient : CSV, ZIP ou
    Excel. La préparation et le filtrage médical/technique sont faits ensuite
    dans utils.cohort.
    """
    if uploaded_file is None:
        return pd.DataFrame()
    return read_uploaded_table(uploaded_file)
